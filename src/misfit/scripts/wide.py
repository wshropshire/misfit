#!/usr/bin/env python3
"""Reshape MISFIT output from long form (one row per gene) to wide form
(one row per assembly, one column per gene).

    # every assembly in one table
    misfit-wide misfit_raw_calls.tsv -o wide.tsv

    # mutation type instead of the HGVS description
    misfit-wide misfit_raw_calls.tsv -o types.tsv \\
        --value Mutation_Type

    # a mixed-organism run split per species, one Excel tab each
    misfit-wide misfit_raw_calls.tsv -o wide.xlsx \\
        --species-map assembly_species_map.tsv

    # the same split, as one TSV per species in a directory
    misfit-wide misfit_raw_calls.tsv -o wide_by_species/ \\
        --species-map assembly_species_map.tsv --format tsv

The species map is the same file `misfit-multi --species-map` takes: a TSV or
CSV with an assembly column (fasta_file / assembly / sample_id) and a species
column (species_wgs / species / organism).
"""
import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from .misfit import load_species_map


def read_long(path, index, columns, value):
    """Read the long-form TSV, checking the three columns we need exist."""
    with open(path, newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter="\t"))
    if not rows:
        raise SystemExit(f"{path} has no data rows")
    have = list(rows[0].keys())
    missing = [c for c in (index, columns, value) if c not in have]
    if missing:
        raise SystemExit(f"{path} has no column(s) {', '.join(missing)}.\n"
                         f"  available: {', '.join(have)}")
    return rows


def pivot(rows, index, columns, value, sep=" | "):
    """Long -> wide. Returns (table, gene_names, n_collapsed).

    A repeated (assembly, gene) pair is joined rather than overwritten: silently
    keeping the last of two disagreeing calls would be the worst option. The
    separator must not be "; ", which MISFIT already uses between variants of a
    single call - two joined calls would be indistinguishable from one call
    listing several variants.
    """
    table = defaultdict(dict)
    genes, collapsed = set(), 0
    for r in rows:
        key, gene = (r[index] or "").strip(), (r[columns] or "").strip()
        if not key or not gene:
            continue
        genes.add(gene)
        val = (r[value] or "").strip()
        if gene in table[key] and table[key][gene] != val:
            table[key][gene] += sep + val
            collapsed += 1
        else:
            table[key][gene] = val
    return table, sorted(genes), collapsed


def to_matrix(table, keys, genes, index, fill, species=None, species_col="Species"):
    """Build the header + rows, species column immediately after the key."""
    header = [index] + ([species_col] if species else []) + genes
    out = [header]
    for k in keys:
        row = [k] + ([species.get(k, "")] if species else [])
        out.append(row + [table[k].get(g, fill) for g in genes])
    return out


def write_tsv(path, matrix):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="") as fh:
        csv.writer(fh, delimiter="\t", lineterminator="\n").writerows(matrix)


_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")


def sheet_name(name, taken):
    """Excel sheet names: <=31 chars, no []:*?/\\, and unique."""
    base = _BAD_SHEET.sub("_", name).strip() or "sheet"
    base = base[:31]
    candidate, n = base, 2
    while candidate.lower() in taken:
        suffix = f"_{n}"
        candidate = base[:31 - len(suffix)] + suffix
        n += 1
    taken.add(candidate.lower())
    return candidate


def write_excel(path, sheets):
    """sheets: list of (name, matrix), written in order."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("writing .xlsx needs openpyxl (pip install openpyxl), "
                         "or use --format tsv")
    wb = Workbook()
    wb.remove(wb.active)
    taken = set()
    for name, matrix in sheets:
        ws = wb.create_sheet(sheet_name(name, taken))
        for row in matrix:
            ws.append(row)
        ws.freeze_panes = "B2"
        for i, head in enumerate(matrix[0], start=1):
            widest = max((len(str(r[i - 1])) for r in matrix[:200]), default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 40)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(
        description="Reshape MISFIT long-form output to wide form.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__.split("\n", 2)[2])
    ap.add_argument("input", help="MISFIT long-form TSV")
    ap.add_argument("-o", "--output", required=True,
                    help="Output file; a directory when --format tsv is combined "
                         "with --species-map")
    ap.add_argument("--index", default="Assembly", help="Row key (default: Assembly)")
    ap.add_argument("--columns", default="Gene",
                    help="Column key (default: Gene)")
    ap.add_argument("--value", default="Mutation_Description",
                    help="Cell value (default: Mutation_Description). "
                         "Mutation_Type, HGVS_c, Notes, Nuc_COV ... all work.")
    ap.add_argument("--species-map", default=None,
                    help="Split the output by species, using the same map "
                         "misfit-multi takes")
    ap.add_argument("--format", choices=["excel", "tsv"], default=None,
                    help="Default: excel when --species-map is given, else tsv")
    ap.add_argument("--fill", default="NA",
                    help="Value for an assembly/gene pair with no row (default: NA)")
    ap.add_argument("--join", default=" | ",
                    help="Separator when one assembly/gene pair has several differing "
                         "rows (default: ' | '). Avoid '; ', which MISFIT uses between "
                         "variants of a single call.")
    ap.add_argument("--species-column", default="Species",
                    help="Name of the added species column (default: Species)")
    args = ap.parse_args()

    fmt = args.format or ("excel" if args.species_map else "tsv")
    rows = read_long(args.input, args.index, args.columns, args.value)
    table, genes, collapsed = pivot(rows, args.index, args.columns, args.value,
                                    sep=args.join)
    keys = sorted(table)
    print(f"{len(rows)} rows -> {len(keys)} {args.index.lower()}(s) x {len(genes)} "
          f"{args.columns.lower()}(s), value = {args.value}")
    if collapsed:
        print(f"  note: {collapsed} repeated {args.index}/{args.columns} pair(s) "
              f"held different values and were joined with '{args.join}'")

    if not args.species_map:
        out = Path(args.output)
        matrix = to_matrix(table, keys, genes, args.index, args.fill)
        if fmt == "excel":
            write_excel(out, [("wide", matrix)])
        else:
            write_tsv(out, matrix)
        print(f"  wrote {out}")
        return

    # ---- split by species
    smap = load_species_map(args.species_map)
    species_of, unmapped = {}, []
    for k in keys:
        sp = smap.get(k) or smap.get(k.split(".")[0])
        if sp:
            species_of[k] = sp
        else:
            unmapped.append(k)
    if unmapped:
        print(f"  warning: {len(unmapped)} assembly(ies) are not in the species map "
              f"and go to 'unmapped': {', '.join(unmapped[:5])}"
              f"{' ...' if len(unmapped) > 5 else ''}")

    groups = defaultdict(list)
    for k in keys:
        groups[species_of.get(k, "unmapped")].append(k)

    # a gene absent from every assembly of a species is dropped from that sheet
    blocks = []
    for sp in sorted(groups, key=lambda s: (s == "unmapped", s)):
        members = groups[sp]
        present = [g for g in genes if any(g in table[k] for k in members)]
        blocks.append((sp, to_matrix(table, members, present, args.index, args.fill,
                                     species=species_of, species_col=args.species_column)))
        print(f"  {sp:38s} {len(members):4d} assembly(ies), {len(present):3d} "
              f"{args.columns.lower()}(s)")

    if fmt == "excel":
        out = Path(args.output)
        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
        write_excel(out, blocks)
        print(f"  wrote {out} ({len(blocks)} tab(s))")
    else:
        outdir = Path(args.output)
        if outdir.suffix.lower() in (".tsv", ".tab", ".txt"):
            outdir = outdir.parent / outdir.stem
        for sp, matrix in blocks:
            safe = re.sub(r"[^A-Za-z0-9._-]+", "_", sp).strip("_") or "unnamed"
            write_tsv(outdir / f"{safe}.tsv", matrix)
        print(f"  wrote {len(blocks)} file(s) in {outdir}/")


if __name__ == "__main__":
    main()
